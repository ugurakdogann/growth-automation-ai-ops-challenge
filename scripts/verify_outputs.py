"""
Verify Outputs — Çıktı Doğrulama Script'i
==========================================

Pipeline çıktısının kalitesini kontrol eder.
Email enrichment metriklerini de raporlar.

Kullanım: python scripts/verify_outputs.py
"""

import csv
import argparse
import sys
from pathlib import Path
from collections import Counter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

PROJECT_ROOT = Path(__file__).parent.parent
DEFAULT_DATA_DIR = PROJECT_ROOT / "data"


def verify(data_dir: Path):
    final_xlsx = data_dir / "leads_final.xlsx"
    email_file = data_dir / "leads_email_enriched.csv"
    rejected_file = data_dir / "rejected_leads.csv"

    print("=" * 60)
    print("  VERIFICATION REPORT")
    print("=" * 60)

    # 1. Ana veri dosyası
    print(f"\n📊 ANA VERİ")
    if not final_xlsx.exists():
        print(f"   ⚠️  {final_xlsx.name} bulunamadı")
    else:
        print(f"   ✅ {final_xlsx.name} mevcut")
        try:
            from openpyxl import load_workbook

            wb = load_workbook(final_xlsx, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            idx = {h: i for i, h in enumerate(headers)}

            def count_nonempty(header):
                if header not in idx:
                    return None
                return sum(1 for row in rows if row[idx[header]])

            email_count = count_nonempty("E-posta")
            estimated_count = count_nonempty("Tahmini E-posta")
            print(f"      Excel satır: {len(rows)}")
            if email_count is not None:
                print(f"      Excel E-posta dolu: {email_count}")
            if estimated_count is not None:
                print(f"      Excel Tahmini E-posta dolu: {estimated_count}")

            removed_detail_columns = [
                "Email Kanit URL",
                "Email Kaynak",
                "Email Pattern",
                "Email Notlari",
                "Email Arama Sorgulari",
                "Email Bulunan URL",
            ]
            leaked_detail_columns = [col for col in removed_detail_columns if col in idx]
            if leaked_detail_columns:
                print(f"      ⚠️  Excel'de teknik email kolonları var: {', '.join(leaked_detail_columns)}")
            else:
                print("      ✅ Excel teknik email kolonları sadeleştirilmiş")

            if "E-posta" in idx and "Email Durumu" in idx:
                xlsx_leaks = [
                    row for row in rows
                    if row[idx["E-posta"]] and row[idx["Email Durumu"]] != "found_public"
                ]
                if xlsx_leaks:
                    print(f"      🚨 Excel email leak: {len(xlsx_leaks)} satır")
                else:
                    print("      ✅ Excel email leak yok")
        except Exception as exc:
            print(f"      ⚠️  Excel detay kontrolü atlandı: {exc}")

    # 2. Email enrichment
    print(f"\n📧 EMAIL ENRICHMENT")
    if not email_file.exists():
        print(f"   ⚠️  {email_file.name} bulunamadı! Email enrichment çalıştırılmamış.")
        print(f"   python scripts/email_enrich.py")
    else:
        with open(email_file, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            leads = list(reader)
            fieldnames = reader.fieldnames or []

        total = len(leads)
        print(f"   ✅ {email_file.name} mevcut — {total} lead")

        # Email status dağılımı
        status_counts = Counter(l.get("email_status", "low_confidence_guess") for l in leads)
        found = status_counts.get("found_public", 0)
        employee_pat = status_counts.get("estimated_from_employee_pattern", 0)
        company_dom = status_counts.get("estimated_from_company_domain", 0)
        low_conf = status_counts.get("low_confidence_guess", 0)
        estimated_total = employee_pat + company_dom + low_conf
        missing = status_counts.get("missing", 0)
        invalid = status_counts.get("invalid", 0)

        print(f"\n   📊 Email Durumu:")
        print(f"      found_public ({95}):                  {found} ({100*found/total:.1f}%)")
        print(f"      estimated_from_employee_pattern ({85}):  {employee_pat} ({100*employee_pat/total:.1f}%)")
        print(f"      estimated_from_company_domain ({70}):    {company_dom} ({100*company_dom/total:.1f}%)")
        print(f"      low_confidence_guess ({50}):             {low_conf} ({100*low_conf/total:.1f}%)")
        print(f"      missing/invalid:                       {missing + invalid}")

        # Confidence dağılımı
        conf_counts = Counter()
        for l in leads:
            try:
                conf_counts[int(l.get("email_confidence", 0))] += 1
            except:
                conf_counts[0] += 1

        print(f"\n   🎯 Email Confidence Dağılımı:")
        for conf in sorted(conf_counts, key=int, reverse=True):
            bar = "█" * (conf_counts[conf] // 5)
            print(f"      {conf}: {conf_counts[conf]:>3} {bar}")

        # Pattern dağılımı
        pat_counts = Counter(l.get("email_pattern", "none") for l in leads if l.get("email_pattern"))
        print(f"\n   📋 Email Pattern Dağılımı:")
        for pat, cnt in pat_counts.most_common():
            print(f"      {pat}: {cnt}")

        # Email'li kayıt oranı
        has_email = sum(1 for l in leads if l.get("email", "").strip())
        has_estimated = sum(1 for l in leads if l.get("estimated_email", "").strip())
        print(f"\n   📈 Email Durumu:")
        print(f"      Gerçek email (email):     {has_email}")
        print(f"      Tahmini email (estimated): {has_estimated}")
        print(f"      Toplam email kapsamı:      {has_email + has_estimated}/{total} "
              f"({100*(has_email+has_estimated)/total:.1f}%)")

        # +++ EVIDENCE_URL CHECK +++
        high_conf_no_evidence = []
        for l in leads:
            try:
                conf = int(l.get("email_confidence", 0))
                if conf >= 70 and not l.get("email_evidence_url", "").strip():
                    high_conf_no_evidence.append(l)
            except: pass
        
        if high_conf_no_evidence:
            print(f"\n   🚨 EVIDENCE EKSİK: {len(high_conf_no_evidence)} lead confidence>=70 ama evidence_url YOK!")
            for l in high_conf_no_evidence[:5]:
                print(f"      {l['full_name']} | conf={l['email_confidence']} | status={l['email_status']}")
        else:
            print(f"\n   ✅ Tüm yüksek confidence tahminlerde evidence_url mevcut")
        
        # +++ EMAIL LEAK CHECK +++
        # Eğer email kolonu dolu ama email_status estimated_pattern ise leak var
        leaked_emails = []
        for l in leads:
            has_real = bool(l.get("email", "").strip())
            status = l.get("email_status", "")
            # found_public dışındaki hiçbir status email kolonuna yazılmamalı
            if has_real and status != "found_public":
                leaked_emails.append(l)
        
        if leaked_emails:
            print(f"\n   🚨 EMAIL LEAK UYARISI: {len(leaked_emails)} lead'de tahmini email, gerçek email kolonuna yazılmış!")
            for l in leaked_emails[:5]:
                print(f"      {l['full_name']} | email=[{l['email']}] | status={status}")
            if len(leaked_emails) > 5:
                print(f"      ... ve {len(leaked_emails)-5} kayıt daha")
        else:
            print(f"\n   ✅ Email leak yok: estimated email'ler gerçek email kolonuna sızmamış")
        
        # Eğer has_email > 0 ama found_public = 0 ise uyarı
        if has_email > 0 and found == 0:
            print(f"   ⚠️  WARNING: {has_email} lead'de E-posta dolu ama found_public=0 — leak olabilir!")
        elif has_email == 0:
            print(f"   ✅ E-posta kolonu boş (beklenen: 0 public email)")
        
        # Şirket bazlı domain dağılımı
        companies_with_pattern = set()
        for l in leads:
            if l.get("email_pattern"):
                companies_with_pattern.add(l.get("company", ""))
        print(f"\n   🏢 Pattern bulunan şirket sayısı: {len(companies_with_pattern)}")

    # 3. Ana lead verisi (leads_enriched.csv)
    enriched_file = data_dir / "leads_enriched.csv"
    if enriched_file.exists():
        with open(enriched_file, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            enriched = list(reader)

        total = len(enriched)
        print(f"\n📊 LEAD ENRICHMENT")

        # Lead score dağılımı
        score_buckets = Counter()
        for l in enriched:
            try:
                s = int(float(l.get("lead_score", 0)))
            except:
                s = 0
            bucket = f"{s//10*10}-{s//10*10+9}"
            score_buckets[bucket] += 1

        print(f"   Lead Score Dağılımı:")
        for bucket in sorted(score_buckets.keys()):
            print(f"      {bucket}: {score_buckets[bucket]}")

        # HR role
        role_cats = Counter(l.get("hr_role", "?") for l in enriched)
        print(f"\n   HR Rol Dağılımı:")
        for cat, cnt in role_cats.most_common():
            print(f"      {cat}: {cnt}")

        # Needs review
        needs_review = sum(1 for l in enriched if l.get("needs_review") == "true")
        print(f"\n   ⚠️  Needs Review: {needs_review}")

        # English need
        eng_counts = Counter()
        for l in enriched:
            try:
                eng_counts[int(l.get("english_need", 0))] += 1
            except:
                pass
        print(f"\n   🇬🇧 English Need:")
        for level in sorted(eng_counts):
            print(f"      Level {level}: {eng_counts[level]}")

        # Top 10 lead score
        try:
            sorted_leads = sorted(enriched, key=lambda l: int(float(l.get("lead_score", "0"))), reverse=True)
            print(f"\n🏆 TOP 10 LEAD:")
            print(f"   {'#':3s} {'İsim':28s} {'Şirket':25s} {'Skor':>5s} {'HR Rol':20s}")
            print(f"   {'-'*85}")
            for i, lead in enumerate(sorted_leads[:10], 1):
                name = lead.get("full_name", "?")[:27]
                company = lead.get("company", "?")[:24]
                score = lead.get("lead_score", "0")
                role = lead.get("hr_role", "?")[:19]
                print(f"   {i:<3d} {name:28s} {company:25s} {score:>5s} {role:20s}")
        except Exception as e:
            print(f"   ⚠️  Sıralama hatası: {e}")

    # 4. Rejected
    if rejected_file.exists():
        with open(rejected_file, "r", encoding="utf-8") as f:
            rejected = list(csv.DictReader(f))
        print(f"\n📋 Rejected: {len(rejected)} kayıt")
    else:
        print(f"\n📋 Rejected: 0")

    # 5. ÖZET
    print(f"\n{'='*60}")
    print(f"  ÖZET")
    print(f"{'='*60}")
    if email_file.exists():
        print(f"  Toplam lead:          {total}")
        print(f"  found_public (95):    {found}")
        print(f"  employee pattern (85):{employee_pat}")
        print(f"  company domain (70):  {company_dom}")
        print(f"  low confidence (50):  {low_conf}")
        print(f"  Tahmini email oranı:  %{100*estimated_total/total:.1f}")
        print(f"  Pattern'li şirket:    {len(companies_with_pattern)}")
        print(f"  Top 10 lead (skor):   hazır")
    print(f"{'='*60}")

    return 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Pipeline output verifier")
    parser.add_argument(
        "--data-dir",
        default=str(DEFAULT_DATA_DIR),
        help="Output folder to verify (default: data/)",
    )
    args = parser.parse_args()
    sys.exit(verify(Path(args.data_dir)))
