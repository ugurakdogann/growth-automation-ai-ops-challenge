"""
Generator v3 — Mesajları Markdown olarak kaydeder.
Her firma için ayrı dosya, CSV'ye mesaj KONULMAZ.
"""

import random
from pathlib import Path
from datetime import datetime


def extract_first_name(full_name: str) -> str:
    if not full_name: return ""
    return full_name.strip().split()[0]


def generate_linkedin_dm(lead: dict) -> str:
    first_name = extract_first_name(lead.get("full_name", ""))
    company = lead.get("company", "şirketiniz")
    pain = lead.get("pain_point", "")
    
    openers = [
        f"Merhaba {first_name},",
        f"{first_name}, iyi günler.",
    ]
    opener = random.choice(openers)
    
    body = f"{company} bünyesindeki İK çalışmalarınızı takip ediyorum. {pain}"
    
    value = ("Konuşarak Öğren olarak kurumlara özel, günde 15 dakikalık "
             "İngilizce konuşma eğitimi sunuyoruz. 500'den fazla kurumsal "
             "müşterimizle çalışanların İngilizce özgüvenini artırıyoruz.")
    
    ctas = [
        "Uygun olursanız 15 dakikalık kısa bir demo planlayabiliriz.",
        "İlginizi çekerse size özel bir çözüm sunumu hazırlayabilirim.",
    ]
    cta = random.choice(ctas)
    
    return f"{opener}\n\n{body}\n\n{value}\n\n{cta}\n\nSaygılarımla."


def generate_cold_email(lead: dict) -> str:
    first_name = extract_first_name(lead.get("full_name", ""))
    company = lead.get("company", "şirketiniz")
    sector = lead.get("sector", "")
    pain = lead.get("pain_point", "")
    
    return f"""Konu: {company} için kurumsal İngilizce eğitim çözümü

Merhaba {first_name},

{company} ekibinin {sector} sektöründeki başarılarını takip ediyoruz.

Konuşarak Öğren olarak Türkiye'nin önde gelen 500'den fazla kurumuna
pratik İngilizce konuşma eğitimi sağlıyoruz.

• Günde 15 dakika — çalışanların iş akışını bölmez
• Ana dili İngilizce olan eğitmenlerle birebir pratik
• Kuruma özel içerik ve ilerleme raporlaması
• Ölçülebilir sonuçlar ve ROI takibi

{pain}

{company} ekibinin İngilizce konuşma becerilerini nasıl
geliştirebileceğimizi konuşmak için 15 dakikalık bir demo
planlamak ister misiniz?

Saygılarımla,
Konuşarak Öğren Kurumsal Ekibi"""


class OutreachGenerator:
    
    def generate(self, leads: list[dict]) -> list[dict]:
        print(f"\n✍️  Generator başladı — {len(leads)} lead")
        for i, lead in enumerate(leads):
            lead["linkedin_dm"] = generate_linkedin_dm(lead)
            lead["cold_email"] = generate_cold_email(lead)
            # email alanına DOKUNMA — enrichment'ten gelen haliyle kalsın
            # estimated_email varsa ayrı kolonda korunur
            lead["generated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if (i + 1) % 50 == 0:
                print(f"  {i+1}/{len(leads)} işlendi...")
        print(f"  ✅ {len(leads)} lead × 2 format")
        return leads
    
    def save_data(self, leads: list[dict], filepath: str):
        """Veriyi formatlı Excel (.xlsx) olarak kaydet."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        filepath = Path(str(filepath).replace(".csv", ".xlsx"))
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        # Mesaj ve teknik email kanıt kolonlarını final Excel'den çıkar.
        # Detaylı kanıtlar leads_email_enriched.csv ve email_enrichment_report.md içinde kalır.
        exclude = {
            "linkedin_dm",
            "cold_email",
            "generated_at",
            "source_url",
            "notes",
            "email_evidence_url",
            "email_pattern_source_note",
            "email_pattern",
            "email_notes",
            "searched_queries",
            "found_urls",
        }
        fieldnames = [k for k in leads[0].keys() if k not in exclude]
        
        # Türkçe başlıklar
        headers_tr = {
            "full_name": "Ad Soyad", "title": "Ünvan", "company": "Şirket",
            "linkedin_url": "LinkedIn", "email": "E-posta", "source": "Kaynak",
            "location": "Konum", "sector": "Sektör", "company_size": "Çalışan",
            "company_focus": "Odak Alanı", "pain_point": "Ağrı Noktası",
            "english_need": "İngilizce İhtiyacı", "outreach_angle": "Yaklaşım",
            "lead_score": "Skor", "hr_role": "HR Rolü",
            "needs_review": "İnceleme",
            "estimated_email": "Tahmini E-posta",
            "email_status": "Email Durumu",
            "email_confidence": "Email Güven Skoru",
        }
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Lead Listesi"
        
        # Başlık satırı
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col, key in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=headers_tr.get(key, key))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Veri satırları
        for row, lead in enumerate(leads, 2):
            for col, key in enumerate(fieldnames, 1):
                val = lead.get(key, "")
                if val is None:
                    val = ""
                # Sayısal değerleri dönüştür
                if key == "lead_score":
                    val = int(val) if val else 0
                elif key == "english_need":
                    val = int(val) if val else 0
                cell = ws.cell(row=row, column=col, value=val)
        
        # Kolon genişliklerini alan adına göre ayarla
        width_map = {
            "full_name": 22, "title": 18, "company": 22, "linkedin_url": 40,
            "email": 28, "source": 12, "location": 18, "sector": 18,
            "company_size": 10, "company_focus": 18, "pain_point": 45,
            "english_need": 5, "outreach_angle": 50, "lead_score": 6,
            "hr_role": 20, "needs_review": 10,
            "estimated_email": 28, "email_status": 16,
            "email_confidence": 14,
        }
        for col, key in enumerate(fieldnames, 1):
            if key in width_map:
                ws.column_dimensions[get_column_letter(col)].width = width_map[key]
        
        # Otomatik filtre
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fieldnames))}{len(leads)+1}"
        
        # Satırları dondur
        ws.freeze_panes = "A2"
        
        wb.save(str(filepath))
        print(f"💾 Veri → {filepath}")
    
    def save_messages_md(self, leads: list[dict], output_dir: str):
        """Mesajları firma bazlı Markdown dosyalarına kaydet."""
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # Firma bazlı grupla
        companies = {}
        for lead in leads:
            company = lead.get("company", "Bilinmeyen")
            if company not in companies:
                companies[company] = []
            companies[company].append(lead)
        
        # Her firma için .md dosyası
        for company, company_leads in companies.items():
            safe_name = "".join(c for c in company if c.isalnum() or c in " _-").strip()
            filepath = output_dir / f"{safe_name}.md"
            
            with open(filepath, "w", encoding="utf-8") as f:
                f.write(f"# {company} — Outreach Mesajları\n\n")
                f.write(f"Toplam: {len(company_leads)} kişi\n\n")
                f.write("---\n\n")
                
                for lead in company_leads:
                    f.write(f"## {lead['full_name']}\n")
                    f.write(f"**{lead.get('title', '')}** | Score: {lead.get('lead_score', '?')}\n\n")
                    
                    if lead.get("linkedin_url"):
                        f.write(f"🔗 {lead['linkedin_url']}\n\n")
                    
                    # Gerçek email (sadece varsa)
                    if lead.get("email", "").strip():
                        f.write(f"📧 {lead['email']}\n\n")
                    
                    # Tahmini email (ayrı etiketle)
                    if lead.get("estimated_email", "").strip():
                        f.write(f"📧 Tahmini email: {lead['estimated_email']}")
                        if lead.get("email_status"):
                            f.write(f" ({lead['email_status']}, güven: {lead.get('email_confidence', '?')})")
                        f.write("\n\n")
                    
                    f.write("### LinkedIn DM\n\n")
                    f.write(f"{lead.get('linkedin_dm', '')}\n\n")
                    
                    f.write("### Cold Email\n\n")
                    f.write(f"{lead.get('cold_email', '')}\n\n")
                    
                    f.write("---\n\n")
        
        print(f"📝 {len(companies)} firma mesajı → {output_dir}/")
    
    def preview(self, lead: dict):
        print(f"\n{'='*60}")
        print(f"  {lead.get('full_name', '?')} — {lead.get('company', '?')}")
        print(f"  Score: {lead.get('lead_score', '?')}")
        if lead.get("email"):
            print(f"  Email: {lead['email']}")
        print(f"{'='*60}")
        print(f"\n📱 DM:\n{lead.get('linkedin_dm', '')[:300]}")
        print(f"\n📧 Email:\n{lead.get('cold_email', '')[:300]}...")
